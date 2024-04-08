# -*- coding: utf-8 -*-
"""
Created on Thu Jul 21 13:55:16 2022

@author: yangxy
"""


import numpy as np
import pandas as pd
import xlwings as xw
import openpyxl as op
import xlwings as xw
import os
import copy

app=xw.App(visible=False,add_book=False)
app.display_alerts=False
app.screen_updating=False

#一、读取各个基金的重仓股是什么
dirpath=r"D:\desktop\mydatabase\fundhold"
dirlist=os.listdir(dirpath)

'''
hang=len(wb.sheets[0].range('A1').current_region.rows)
lie=len(wb.sheets[0].range('A1').current_region.columns)
df=wb.sheets[0].range((1,1),(hang,lie)).options(pd.DataFrame,index=False).value
'''

greatlis=dict()
ok=dict()
for filename in dirlist:
    print(filename)
    if os.path.isdir(os.path.join(dirpath,filename)):
        continue
    elif  filename=="ok.xlsx":
        wb=app.books.open(os.path.join(dirpath,filename))
        wb.sheets[0].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
        ok["重仓股代码"]=df
        wb.close()
        continue
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    greatdf=dict()
    for she in range(len(wb.sheets)):
        wb.sheets[she].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[she].range('A1').options(pd.DataFrame,index=False,expand="table").value
        greatdf[wb.sheets[she].name]=df
    greatlis[filename]=greatdf
    wb.close()

tmp1=set()
for i in greatlis.keys():
    tmp1=set(greatlis[i].keys()) | tmp1
shtlis=copy.deepcopy(tmp1)
del tmp1
holddic=dict()
for sht in shtlis:
    #break
    df=pd.DataFrame(columns=["date"])
    for sce in greatlis.values():
        #break
        #greatlis.keys()
        if sht in sce.keys():
            tmp1=copy.deepcopy(sce[sht])
            tmp1.index=pd.to_datetime(tmp1['date'])
            tmp1=tmp1.resample('3M',axis=0,closed="right",label="right").last()
            tmp1['date']=tmp1.index
            tmp1.index.name='index'
            df=pd.merge(df,tmp1,how="outer",on="date")
            tmp1=df.copy()
            i=sht
            tmp2=pd.Series(tmp1.columns).apply(lambda x:True if x[-2:-1]=="_" else False)
            if len(tmp2[tmp2])>0:
                tmp3=tmp1.loc[:,tmp1.columns[tmp2]]
                tmp4=pd.Series(tmp3.columns).apply(lambda x:x[:-2])
                tmp4=set(tmp4)
                tmp7=pd.DataFrame(index=tmp1.index)
                for j in tmp4:
                    #break
                    tmp5=tmp3.loc[:,j+"_x"]
                    tmp6=tmp3.loc[:,j+"_y"]
                    tmp5=tmp5.replace(0,np.nan)
                    tmp6=tmp6.replace(0,np.nan)
                    for k in range(len(tmp5)):
                        #break
                        if (tmp5[k]==tmp5[k]) or (tmp5[k]==""):
                            pass
                        else:
                            tmp5[k]=tmp6[k]
                    tmp5=pd.DataFrame(tmp5,index=tmp1.index)
                    tmp5.columns=[j]
                    tmp7=pd.concat([tmp7,tmp5],axis=1)
                tmp1=tmp1.loc[:,tmp1.columns[tmp2.apply(lambda x:not(x))]]
                del tmp2,tmp3,tmp4,tmp5,tmp6
                tmp7=pd.concat([tmp1['date'],tmp7],axis=1)
                tmp1=pd.concat([tmp7,tmp1.iloc[:,1:]],axis=1)
                df=tmp1.copy()
        else:
            pass
    df=df.sort_values(by='date',axis=0,ascending=True,inplace=False,na_position='last')
    df=df.reset_index(drop=True)
    holddic[sht]=df
del sht,sce,tmp1
for k in holddic.keys():
    #break
    tmp1=holddic[k]
    holddic[k]=tmp1.applymap(
        lambda x:None if (x==0.0) or (x!=x) or (x=="") or (x is None) else str(x)
        )
tmp2,tmp7,j=0,0,0
del greatdf,greatlis,i,k,she,shtlis,tmp1,tmp2,tmp7,j,df,dirlist,dirpath,filename,wb


#二、读取各个基金的重仓股的占比
dirpath=r"D:\desktop\mydatabase\fundholdpct"
dirlist=os.listdir(dirpath)

'''
hang=len(wb.sheets[0].range('A1').current_region.rows)
lie=len(wb.sheets[0].range('A1').current_region.columns)
df=wb.sheets[0].range((1,1),(hang,lie)).options(pd.DataFrame,index=False).value
'''
greatlis=dict()
for filename in dirlist:
    print(filename)
    if os.path.isdir(os.path.join(dirpath,filename)):
        continue
    elif  filename=="ok.xlsx":
        wb=app.books.open(os.path.join(dirpath,filename))
        wb.sheets[0].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[0].range('A1').options(pd.DataFrame,index=False,expand="table").value
        ok["重仓股占比"]=df
        continue
    #break
    wb=app.books.open(os.path.join(dirpath,filename))
    #wb[wb.sheetnames[0]].title
    greatdf=dict()
    for she in range(len(wb.sheets)):
        wb.sheets[she].name
        wb.sheets[0].used_range.last_cell.row
        wb.sheets[0].used_range.last_cell.column
        df=wb.sheets[she].range('A1').options(pd.DataFrame,index=False,expand="table").value
        greatdf[wb.sheets[she].name]=df
    greatlis[filename]=greatdf
    wb.close()

tmp1=set()
for i in greatlis.keys():
    tmp1=set(greatlis[i].keys()) | tmp1
shtlis=copy.deepcopy(tmp1)
del tmp1
holdpctdic=dict()
for sht in shtlis:
    #break
    df=pd.DataFrame(columns=["date"])
    for sce in greatlis.values():
        #break
        #greatlis.keys()
        if sht in sce.keys():
            tmp1=copy.deepcopy(sce[sht])
            tmp1.index=pd.to_datetime(tmp1['date'])
            tmp1=tmp1.resample('3M',axis=0,closed="right",label="right").last()
            tmp1['date']=tmp1.index
            tmp1.index.name='index'
            df=pd.merge(df,tmp1,how="outer",on="date")
            tmp1=df.copy()
            i=sht
            tmp2=pd.Series(tmp1.columns).apply(lambda x:True if x[-2:-1]=="_" else False)
            if len(tmp2[tmp2])>0:
                tmp3=tmp1.loc[:,tmp1.columns[tmp2]]
                tmp4=pd.Series(tmp3.columns).apply(lambda x:x[:-2])
                tmp4=set(tmp4)
                tmp7=pd.DataFrame(index=tmp1.index)
                for j in tmp4:
                    #break
                    tmp5=tmp3.loc[:,j+"_x"]
                    tmp6=tmp3.loc[:,j+"_y"]
                    tmp5=tmp5.replace(0,np.nan)
                    tmp6=tmp6.replace(0,np.nan)
                    for k in range(len(tmp5)):
                        #break
                        if tmp5[k]==tmp5[k] or (tmp5[k]==""):
                            pass
                        else:
                            tmp5[k]=tmp6[k]
                    tmp5=pd.DataFrame(tmp5,index=tmp1.index)
                    tmp5.columns=[j]
                    tmp7=pd.concat([tmp7,tmp5],axis=1)
                tmp1=tmp1.loc[:,tmp1.columns[tmp2.apply(lambda x:not(x))]]
                del tmp2,tmp3,tmp4,tmp5,tmp6
                tmp7=pd.concat([tmp1['date'],tmp7],axis=1)
                tmp1=pd.concat([tmp7,tmp1.iloc[:,1:]],axis=1)
                df=tmp1.copy()
        else:
            pass
    df=df.sort_values(by='date',axis=0,ascending=True,inplace=False,na_position='last')
    df=df.reset_index(drop=True)
    holdpctdic[sht]=df
del sht,sce,tmp1
for k in holdpctdic.keys():
    #break
    tmp1=holdpctdic[k]
    holdpctdic[k]=tmp1.applymap(
        lambda x:None if (x==0.0) or (x!=x) or (x=="") or (x is None) else x
        )
tmp2,tmp7,j=0,0,0
del greatdf,greatlis,i,k,she,shtlis,tmp1,tmp2,tmp7,j,dirpath,dirlist,df,filename,wb

#funds=pd.DataFrame(zcrateshtlis[0].iloc[0,1:].copy())
#funds.index=range(100)

#三、读取各个基金的所属行业
wb=op.load_workbook(r'D:\desktop\indus_rolling\code2indus.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
code2indus=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        code2indus.iat[i,j]=she.cell(i+1,j+1).value
wb.close()

#读取各个基金所属行业（港股）
wb=op.load_workbook(r'D:\desktop\indus_rolling\code2indushk.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
code2indushk=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        code2indushk.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
'''
sheet=wb[wb.sheetnames[0]]
sheet.cell(1,2).value
#wb[wb.sheetnames[0]]['B2'].value
sheet.max_row
sheet.max_column
for i in wb[wb.sheetnames[0]]:
    print(i)
'''

code2indus=code2indus.drop([0],axis=0)
code2indushk=code2indushk.drop([0],axis=0)
code2indus=pd.concat([code2indus,code2indushk],axis=0)
code2indus.columns=['code','industry']
code2indus.index=range(len(code2indus.index))
del code2indushk

wb=op.load_workbook(r'D:\desktop\indus_rolling\indusname.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
indusname=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        indusname.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
indusname=indusname.drop([0],axis=0)
indusname.columns=['induscode','name']

wb=op.load_workbook(r'D:\desktop\indus_rolling\indusret.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
indusret=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        indusret.iat[i,j]=she.cell(i+1,j+1).value
wb.close()

#indusret.iloc[2:,1:].rank(axis=1,method="average",ascending=False)

#四、计算各个基金每个行业占基金股票投资的比重
code2indus0=code2indus.applymap(lambda x: x[:len(x)-3] if not(x is None) else None)
indusret.columns=indusret.iloc[0,:]
indusret.drop([indusret.index[0]],axis=0,inplace=True)
indusret.index=indusret.iloc[:,0]
indusret.drop([indusret.index[0]],axis=0,inplace=True)
indusret.drop([indusret.columns[0]],axis=1,inplace=True)
indusretrank=indusret.rank(axis=1,method="average",ascending=False)
del i,j,she,wb
'''
for diji in range(10):
    sh=holddic["第"+str(diji+1)+"大重仓股代码"].copy()
    sh.columns=sh.iloc[0,:]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.index=sh.iloc[:,0]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.drop([sh.columns[0]],axis=1,inplace=True)
    zcshtlis[diji]=sh.copy()
    
    sh=zcrateshtlis[diji].copy()
    sh.columns=sh.iloc[0,:]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.index=sh.iloc[:,0]
    sh.drop([sh.index[0]],axis=0,inplace=True)
    sh.drop([sh.columns[0]],axis=1,inplace=True)
    zcrateshtlis[diji]=sh.copy()
'''

#五、检查源数据的日期是否有缺角,并确定最终时间范围
tmp1=0
tmp2=set()
for i in holddic.keys():
    print(len(set(holddic[i]['date'])))
    if tmp1!=len(set(holddic[i]['date'])) and (tmp1!=0):
        print("Wrong!")
    tmp1=len(set(holddic[i]['date']))
    tmp2=set(pd.to_datetime(holddic[i]['date'])) | tmp2
for i in holdpctdic.keys():
    print(len(set(holdpctdic[i]['date'])))
    if tmp1!=len(set(holdpctdic[i]['date'])) and (tmp1!=0):
        print("Wrong!")
    tmp1=len(set(holdpctdic[i]['date']))
for i in tmp2:
    if pd.to_datetime(i) in set(indusret.index):
        pass
    else:
        print("Wrong!")
        print(i)
tmp2=set(indusret.index) & tmp2
datelis=pd.Series(list(tmp2)).sort_values(ascending=True).reset_index(drop=True)
datelis.name="date"

#六、检查源数据的基金池是否有缺角,并确定最终基金范围
tmp1=0
tmp2=set()
flag=0
for i in ok.values():
    flag+=1
    if flag==1:
        tmp2=set(i["fundname"])
    else:
        tmp2=set(i["fundname"]) & tmp2
fundlis=pd.Series(list(tmp2)).sort_values(ascending=True).reset_index(drop=True)
fundlis.name="fund"
del tmp1,tmp2,ok,i,flag

defen=pd.DataFrame(index=datelis,columns=fundlis)
induschange=defen.copy()
daibiao=defen.copy()
for i in range(len(fundlis.index)):
    lagspread=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
    for dateindex in range(len(datelis)):
        induspack=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
        indusspread=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
        for diji in range(len(holddic.keys())):
            "第"+str(diji+1)+"大重仓股代码"
            tmp1=holddic["第"+str(diji+1)+"大重仓股代码"]
            tmp1['date']=pd.to_datetime(tmp1['date'])
            tmp2=tmp1[tmp1["date"]==datelis[dateindex]].loc[:,fundlis[i]].iat[0]
            if tmp2==None or tmp2!=tmp2:
                continue
            if len(code2indus0[code2indus0['code']==tmp2].index)==0:
                continue
            hangye=code2indus0[code2indus0['code']==tmp2].iat[0,1]
            hangye=hangye+".SI"
            "第"+str(diji+1)+"大重仓股占比"
            tmp1=holdpctdic["第"+str(diji+1)+"大重仓股占比"]
            tmp2=tmp1[tmp1["date"]==datelis[dateindex]].loc[:,fundlis[i]].iat[0]
            indusspread.loc[hangye,:][0]=indusspread.loc[hangye,][0]+tmp2
            induspack.loc[hangye,:][0]=(
                induspack.loc[hangye,][0]+
                indusretrank.loc[indusret.index[dateindex],hangye]*tmp2
                )
        if dateindex==0:
            induschange.loc[datelis[dateindex],fundlis[i]]=None
        else:
            induschange.loc[datelis[dateindex],fundlis[i]]=((lagspread-indusspread)**2).sum()[0]
        lagspread=indusspread.copy()
        defen.loc[datelis[dateindex],fundlis[i]]=induspack.sum()[0]

defen.to_excel(r'D:\desktop\indus_rolling\defennew.xlsx',header=True,index=True)
induschange.to_excel(r'D:\desktop\indus_rolling\induschangenew.xlsx',header=True,index=True)

indusrate_ex=pd.DataFrame(0.00,index=indusret.columns,columns=['rate'])
for i in range(len(fundlis.index)):
    for dateindex in range(len(datelis)):
        indusrate_now=pd.DataFrame(0.00,index=indusret.columns,columns=['rate']).copy()
        for diji in range(len(holddic.keys())):
            "第"+str(diji+1)+"大重仓股代码"
            tmp1=holddic["第"+str(diji+1)+"大重仓股代码"]
            tmp2=tmp1[tmp1["date"]==datelis[dateindex]].loc[:,fundlis[i]].iat[0]
            if tmp2==None or tmp2!=tmp2:
                continue
            if len(code2indus0[code2indus0['code']==tmp2].index)==0:
                continue
            hangye=code2indus0[code2indus0['code']==tmp2].iat[0,1]
            hangye=hangye+".SI"
            "第"+str(diji+1)+"大重仓股占比"
            tmp1=holdpctdic["第"+str(diji+1)+"大重仓股占比"]
            tmp2=tmp1[tmp1["date"]==datelis[dateindex]].loc[:,fundlis[i]].iat[0]
            indusrate_now.loc[hangye,:][0]=indusrate_now.loc[hangye,:][0]+tmp2
        zuidanew=indusrate_now[indusrate_now['rate']==indusrate_now.max()[0]].index[0] if indusrate_now.max()[0]>0 else None
        zuidaold=indusrate_ex[indusrate_ex['rate']==indusrate_ex.max()[0]].index[0] if indusrate_ex.max()[0]>0 else None
        if (zuidanew is None) or (zuidaold is None):
            pass
        elif zuidanew!=zuidaold:
            huancangstr=indusrate_ex[indusrate_ex['rate']==indusrate_ex.max()[0]].index[0]
            huancangstr=indusname[indusname['induscode']==huancangstr].iat[0,1]
            miaoshu="由重仓"+huancangstr+"行业调仓为"
            huancangstr=indusrate_now[indusrate_now['rate']==indusrate_now.max()[0]].index[0]
            huancangstr=indusname[indusname['induscode']==huancangstr].iat[0,1]
            miaoshu=miaoshu+huancangstr+"行业"
            daibiao.iat[dateindex,i]=miaoshu
        indusrate_ex=indusrate_now.copy()

daibiao.to_excel(r'D:\desktop\indus_rolling\daibiaonew.xlsx',header=True,index=True)

#======================================================
#七、程序外整理defennew后，后续调试可从此开始
import numpy as np
import pandas as pd
import xlwings as xw
import openpyxl as op
import xlwings as xw
import os
import copy


app=xw.App(visible=False,add_book=False)
app.display_alerts=False
app.screen_updating=False


wb=op.load_workbook(r'D:\desktop\indus_rolling\defen.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
defen=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        defen.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
sh=defen.copy()
sh.columns=sh.iloc[0,:]
sh.drop([sh.index[0]],axis=0,inplace=True)
sh.index=sh.iloc[:,0]
sh.drop([sh.columns[0]],axis=1,inplace=True)
defen=sh.copy()

wb=op.load_workbook(r'D:\desktop\indus_rolling\daibiao.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
daibiao=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        daibiao.iat[i,j]=she.cell(i+1,j+1).value
wb.close()
sh=daibiao.copy()
sh.columns=sh.iloc[0,:]
sh.drop([sh.index[0]],axis=0,inplace=True)
sh.index=sh.iloc[:,0]
sh.drop([sh.columns[0]],axis=1,inplace=True)
daibiao=sh.copy()

wb=op.load_workbook(r'D:\desktop\indus_rolling\induschange.xlsx',data_only=True)
she=wb[wb.sheetnames[0]]
induschange=pd.DataFrame(None,index=range(she.max_row),columns=range(she.max_column))
for i in range(she.max_row):
    for j in range(she.max_column):
        induschange.iat[i,j]=she.cell(i+1,j+1).value
sh=induschange.copy()
sh.columns=sh.iloc[0,:]
sh.drop([sh.index[0]],axis=0,inplace=True)
sh.index=sh.iloc[:,0]
sh.drop([sh.columns[0]],axis=1,inplace=True)
induschange=sh.copy()
induschange=induschange.applymap(lambda x:None if x==0 else x)
del i,j,sh,she,wb
defen.replace([0,None],value=np.nan,inplace=True)
daibiao.replace([0,None],value=np.nan,inplace=True)
induschange.replace([0,None],value=np.nan,inplace=True)

#八、给不同年份一个权重，将defen和change历年数据加权综合成一个数
xrank=defen.rank(
    axis=1,method='average',numeric_only=None,
    na_option='keep',ascending=True,pct=True
    )
wet=pd.Series(0.0,index=defen.index)
wet.name="weight"
for i in reversed(range(len(wet))):
    tmp1=1.05-((len(wet)-i-1) // 2 )*0.05
    wet.iat[i]=tmp1
huizong=pd.DataFrame(0.0,index=defen.columns,columns=['Rx','Rchange',"N"])
for i in xrank.columns:
    tmp1=xrank[i][xrank[i]==xrank[i]].copy()#把空值剔除了
    tmp2=wet.loc[tmp1.index].copy()
    tmp3=(tmp1*tmp2).sum() / tmp2.sum()
    huizong.loc[i,"Rx"]=tmp3
    huizong.loc[i,"N"]=len(tmp1)
changerank=induschange.rank(
    axis=1,method='average',numeric_only=None,
    na_option='keep',ascending=True,pct=True
    )
for i in changerank.columns:
    tmp1=changerank[i][changerank[i]==changerank[i]].copy()#把空值剔除了
    tmp2=wet.loc[tmp1.index].copy()
    tmp3=(tmp1*tmp2).sum() / tmp2.sum()
    huizong.loc[i,"Rchange"]=tmp3
huizong["总得分"]=huizong["Rx"]*0.7+huizong["Rchange"]*0.3
for i in daibiao.columns:
    lis=list()
    for j in daibiao[i].index:
        tmp1=daibiao.loc[j,i]
        if tmp1==tmp1:
            tmp2=j.strftime("%Y%m%d")
            lis.append(tmp2+str(tmp1)+"\n")
    huizong.loc[i,"经典战役"]="".join(lis)
huizong.to_excel(r'D:\desktop\indus_rolling\huizongnew.xlsx',header=True,index=True)

app.kill()

